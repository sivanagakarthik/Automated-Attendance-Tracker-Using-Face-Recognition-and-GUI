import cv2
import os
import numpy as np
from tkinter import *
from tkinter import ttk, simpledialog, messagebox, filedialog
from PIL import Image, ImageTk
import datetime
import csv
import hashlib
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table as RLTable, TableStyle

# -------------------- Constants --------------------
IMAGES_PER_STUDENT = 30                   # Paper §III-D: 30 images per student
FACE_WIDTH, FACE_HEIGHT = 200, 200
RECOGNITION_CONFIDENCE_THRESHOLD = 75     # Paper §IV-B: valid below 70-80
CASCADE_PATH = "haarcascade_frontalface_default.xml"
DATASET_DIR = "dataset"
TRAINER_FILE = "trainer.yml"
LABELS_FILE = "labels.txt"
ATTENDANCE_FILE = "attendance.csv"
USERS_FILE = "users.csv"
UNKNOWN_LOG_FILE = "unknown_faces.csv"
SAVED_SESSIONS_FILE = "saved_sessions.json"

# -------------------- Ensure Haar Cascade --------------------
if not os.path.exists(CASCADE_PATH):
    import urllib.request
    url = "https://raw.githubusercontent.com/opencv/opencv/master/data/haarcascades/haarcascade_frontalface_default.xml"
    urllib.request.urlretrieve(url, CASCADE_PATH)
face_cascade = cv2.CascadeClassifier(CASCADE_PATH)

# -------------------- Users --------------------
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def ensure_users_file():
    if not os.path.exists(USERS_FILE) or os.path.getsize(USERS_FILE) == 0:
        with open(USERS_FILE, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["teacher", hash_password("1234")])

def load_users():
    users = {}
    with open(USERS_FILE, "r") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) == 2:
                users[row[0]] = row[1]
    return users

def save_users(users_dict):
    with open(USERS_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        for username, pwd_hash in users_dict.items():
            writer.writerow([username, pwd_hash])

ensure_users_file()

# -------------------- Attendance --------------------
def mark_attendance_once(student_id, name):
    """Mark attendance only once per student per day. Returns True if newly marked."""
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    if not os.path.exists(ATTENDANCE_FILE):
        with open(ATTENDANCE_FILE, "w", newline="") as f:
            csv.writer(f).writerow(["ID", "Name", "Date", "Time", "Status"])
    with open(ATTENDANCE_FILE, "r", newline="") as f:
        for row in csv.reader(f):
            if not row or row[0].strip().upper() in ("ID", ""):
                continue  # skip header
            if len(row) >= 3 and row[0] == str(student_id) and row[2] == today:
                return False
    now = datetime.datetime.now()
    with open(ATTENDANCE_FILE, "a", newline="") as f:
        csv.writer(f).writerow([
            student_id, name,
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M:%S"),
            "Present"
        ])
    return True

def log_unknown_face():
    """Log unknown face detection for audit purposes."""
    if not os.path.exists(UNKNOWN_LOG_FILE):
        with open(UNKNOWN_LOG_FILE, "w", newline="") as f:
            csv.writer(f).writerow(["Date", "Time"])
    now = datetime.datetime.now()
    with open(UNKNOWN_LOG_FILE, "a", newline="") as f:
        csv.writer(f).writerow([now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")])

# -------------------- Preprocessing --------------------
def preprocess_face(face_gray):
    """
    Paper §IV-I: Apply CLAHE (adaptive histogram equalization) to improve
    recognition under varied lighting conditions.
    """
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    return clahe.apply(face_gray)

# ==================== THEME CONSTANTS ====================
BG_DARK      = "#0d1117"   # deep space black
BG_CARD      = "#161b22"   # card surface
BG_SIDEBAR   = "#0d1117"   # sidebar bg
ACCENT       = "#00d4aa"   # teal glow
ACCENT2      = "#00a8ff"   # blue accent
ACCENT_RED   = "#ff5f6d"   # danger / unknown
TEXT_PRIMARY = "#e6edf3"   # bright text
TEXT_DIM     = "#8b949e"   # muted text
BORDER       = "#30363d"   # subtle border
BTN_HOVER    = "#1f6feb"   # button hover
SUCCESS_BG   = "#0d2818"
ERROR_BG     = "#2d1215"

FONT_TITLE   = ("Helvetica", 22, "bold")
FONT_HEADING = ("Helvetica", 13, "bold")
FONT_LABEL   = ("Helvetica", 10)
FONT_MONO    = ("Courier", 11)
FONT_SMALL   = ("Helvetica", 9)

# -------------------- Login Window --------------------
class LoginWindow:
    def __init__(self, master, on_success):
        self.master = master
        self.master.title("FaceID Attendance — Login")
        self.master.geometry("480x520")
        self.master.configure(bg=BG_DARK)
        self.master.resizable(False, False)
        self.on_success = on_success
        self.users = load_users()
        self._pulse_state = 0
        self._build_ui()
        self._animate_pulse()

    def _build_ui(self):
        m = self.master

        # ── top glow bar ──
        glow = Canvas(m, height=4, bg=BG_DARK, highlightthickness=0)
        glow.pack(fill=X)
        glow.create_rectangle(0, 0, 480, 4, fill=ACCENT, outline="")

        # ── icon + title ──
        top = Frame(m, bg=BG_DARK)
        top.pack(pady=(32, 8))

        # animated pulsing circle behind icon
        self._pulse_canvas = Canvas(top, width=100, height=100,
                                    bg=BG_DARK, highlightthickness=0)
        self._pulse_canvas.pack()
        self._pulse_canvas.create_oval(10, 10, 90, 90, fill=BG_CARD,
                                       outline=ACCENT, width=2, tags="circle")
        self._pulse_canvas.create_text(50, 50, text="👤",
                                       font=("Helvetica", 28), fill=TEXT_PRIMARY, tags="icon")

        Label(m, text="FaceID Attendance", font=FONT_TITLE,
              bg=BG_DARK, fg=TEXT_PRIMARY).pack()
        Label(m, text="Real-Time Face Recognition Attendance System",
              font=FONT_SMALL, bg=BG_DARK, fg=TEXT_DIM).pack(pady=(2, 0))

        # ── card ──
        card = Frame(m, bg=BG_CARD, bd=0, relief="flat",
                     highlightthickness=1, highlightbackground=BORDER)
        card.pack(padx=50, pady=24, fill=X, ipady=10)

        Label(card, text="A D M I N I S T R A T O R   L O G I N",
              font=("Helvetica", 9, "bold"),
              bg=BG_CARD, fg=ACCENT).pack(pady=(18, 12))

        # username
        uf = Frame(card, bg=BG_CARD)
        uf.pack(fill=X, padx=28, pady=6)
        Label(uf, text="USERNAME", font=("Helvetica", 8, "bold"),
              bg=BG_CARD, fg=TEXT_DIM).pack(anchor="w")
        self.username_entry = Entry(uf, font=("Helvetica", 12),
                                    bg="#21262d", fg=TEXT_PRIMARY, bd=0,
                                    insertbackground=ACCENT, relief="flat",
                                    highlightthickness=1, highlightbackground=BORDER,
                                    highlightcolor=ACCENT)
        self.username_entry.pack(fill=X, ipady=8, pady=(2, 0))

        # password
        pf = Frame(card, bg=BG_CARD)
        pf.pack(fill=X, padx=28, pady=6)
        Label(pf, text="PASSWORD", font=("Helvetica", 8, "bold"),
              bg=BG_CARD, fg=TEXT_DIM).pack(anchor="w")
        self.password_entry = Entry(pf, show="●", font=("Helvetica", 12),
                                    bg="#21262d", fg=TEXT_PRIMARY, bd=0,
                                    insertbackground=ACCENT, relief="flat",
                                    highlightthickness=1, highlightbackground=BORDER,
                                    highlightcolor=ACCENT)
        self.password_entry.pack(fill=X, ipady=8, pady=(2, 0))
        self.password_entry.bind("<Return>", lambda e: self.check_login())

        # error label
        self.err_label = Label(card, text="", font=FONT_SMALL,
                               bg=BG_CARD, fg=ACCENT_RED)
        self.err_label.pack(pady=(4, 0))

        # login button
        self.login_btn = Button(card, text="SIGN IN  →",
                                font=("Helvetica", 11, "bold"),
                                bg=ACCENT, fg=BG_DARK, bd=0, relief="flat",
                                activebackground="#00b894", activeforeground=BG_DARK,
                                cursor="hand2", command=self.check_login)
        self.login_btn.pack(padx=28, pady=(10, 20), fill=X, ipady=10)
        self.login_btn.bind("<Enter>", lambda e: self.login_btn.config(bg="#00b894"))
        self.login_btn.bind("<Leave>", lambda e: self.login_btn.config(bg=ACCENT))

        Label(m, text="Default credentials: teacher / 1234",
              font=FONT_SMALL, bg=BG_DARK, fg=TEXT_DIM).pack()

        self.username_entry.focus_set()

    def _animate_pulse(self):
        """Gently pulse the icon circle."""
        self._pulse_state = (self._pulse_state + 1) % 60
        alpha = 0.4 + 0.6 * abs(self._pulse_state - 30) / 30
        # interpolate teal brightness
        r = int(0 * alpha); g = int(212 * alpha); b = int(170 * alpha)
        color = f"#{r:02x}{g:02x}{b:02x}"
        self._pulse_canvas.itemconfig("circle", outline=color)
        self.master.after(50, self._animate_pulse)

    def check_login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get()
        if username in self.users and self.users[username] == hash_password(password):
            self.master.destroy()
            self.on_success()
        else:
            self.err_label.config(text="⚠  Invalid username or password")
            self.password_entry.delete(0, END)
            self.password_entry.focus_set()

# -------------------- Notification Bar --------------------
class NotificationBar:
    """Sleek notification toast at the bottom of the main panel."""
    def __init__(self, parent):
        self.label = Label(parent, text="", font=("Helvetica", 10),
                           bg=BG_DARK, fg=TEXT_DIM, anchor="w", padx=14,
                           relief="flat")
        self.label.pack(fill=X, side=BOTTOM, ipady=6)
        self._after_id = None

    def show(self, message, color=ACCENT, duration=4000):
        if self._after_id:
            self.label.after_cancel(self._after_id)
        self.label.config(text=f"  ●  {message}", fg=color,
                          bg=SUCCESS_BG if color == ACCENT else ERROR_BG)
        self._after_id = self.label.after(duration,
                            lambda: self.label.config(text="", bg=BG_DARK))

    def error(self, message):
        self.show(message, color=ACCENT_RED, duration=5000)

# -------------------- Attendance App --------------------
class AttendanceApp:
    def __init__(self, master):
        self.master = master
        master.title("FaceID Attendance System")
        master.geometry("1200x720")
        master.configure(bg=BG_DARK)
        master.minsize(960, 620)

        self.labels_map = self.load_labels()
        self.recognizer = cv2.face.LBPHFaceRecognizer_create()
        if os.path.exists(TRAINER_FILE):
            self.recognizer.read(TRAINER_FILE)

        self.running = False
        self.cap = None
        self._unknown_log_cooldown = {}
        self._session_logged = set()     # tracks who already got a log entry this session
        self.filter_date = None
        self._unknown_count = 0
        self._cam_border_color = BORDER
        self._cam_anim_step = 0
        self._log_saved = True           # True = current log has been saved
        self._saved_sessions = self._load_sessions_from_disk()  # persisted across logins

        self._apply_ttk_style()
        self.setup_gui()
        self.update_clock()
        self._animate_cam_border()

    def _apply_ttk_style(self):
        """Apply dark theme to ttk widgets."""
        style = ttk.Style()
        style.theme_use("default")
        style.configure("Dark.Treeview",
                        background=BG_CARD,
                        foreground=TEXT_PRIMARY,
                        fieldbackground=BG_CARD,
                        rowheight=28,
                        font=("Helvetica", 10),
                        borderwidth=0)
        style.configure("Dark.Treeview.Heading",
                        background="#21262d",
                        foreground=ACCENT,
                        font=("Helvetica", 9, "bold"),
                        relief="flat",
                        borderwidth=0)
        style.map("Dark.Treeview",
                  background=[("selected", "#1f3a4d")],
                  foreground=[("selected", TEXT_PRIMARY)])
        style.configure("Dark.Vertical.TScrollbar",
                        background="#21262d",
                        troughcolor=BG_CARD,
                        borderwidth=0,
                        arrowcolor=TEXT_DIM)

    # ------------------------------------------------------------------ GUI
    def setup_gui(self):
        # ══════════════════ SIDEBAR ══════════════════
        sidebar = Frame(self.master, bg=BG_SIDEBAR, width=240)
        sidebar.pack(side=LEFT, fill=Y)
        sidebar.pack_propagate(False)

        # ── Fixed top: accent bar + logo + subtitle ──
        top_fixed = Frame(sidebar, bg=BG_SIDEBAR)
        top_fixed.pack(fill=X)

        Canvas(top_fixed, height=3, bg=BG_SIDEBAR, highlightthickness=0).pack(fill=X)
        Frame(top_fixed, bg=ACCENT, height=3).pack(fill=X)

        logo_frame = Frame(top_fixed, bg=BG_SIDEBAR)
        logo_frame.pack(fill=X, pady=(20, 4), padx=16)
        Label(logo_frame, text="◈", font=("Helvetica", 26),
              bg=BG_SIDEBAR, fg=ACCENT).pack(side=LEFT)
        name_frame = Frame(logo_frame, bg=BG_SIDEBAR)
        name_frame.pack(side=LEFT, padx=8)
        Label(name_frame, text="FaceID", font=("Helvetica", 16, "bold"),
              bg=BG_SIDEBAR, fg=TEXT_PRIMARY).pack(anchor="w")
        Label(name_frame, text="Attendance System", font=("Helvetica", 8),
              bg=BG_SIDEBAR, fg=TEXT_DIM).pack(anchor="w")

        self.header_label = Label(top_fixed, text="",
                                  font=("Helvetica", 9, "italic"),
                                  bg=BG_SIDEBAR, fg=ACCENT, wraplength=210)
        self.header_label.pack(padx=16, pady=(0, 8))
        self.typewriter_text("Real-Time Face Recognition  •  LBPH Algorithm")

        Frame(top_fixed, bg=BORDER, height=1).pack(fill=X, padx=16, pady=4)

        # ── Fixed bottom: clock ──
        bottom_fixed = Frame(sidebar, bg=BG_SIDEBAR)
        bottom_fixed.pack(side=BOTTOM, fill=X)
        Label(bottom_fixed, text="Face Recognition Attendance",
              font=("Helvetica", 7), bg=BG_SIDEBAR, fg="#444").pack(pady=(0, 6))
        self.clock_label = Label(bottom_fixed, font=FONT_MONO,
                                 bg=BG_SIDEBAR, fg=TEXT_DIM, justify=CENTER)
        self.clock_label.pack(pady=(6, 0))
        Frame(bottom_fixed, bg=BORDER, height=1).pack(fill=X, padx=16, pady=4)

        # ── Scrollable middle: all nav buttons ──
        scroll_outer = Frame(sidebar, bg=BG_SIDEBAR)
        scroll_outer.pack(fill=BOTH, expand=True)

        _sb_canvas = Canvas(scroll_outer, bg=BG_SIDEBAR,
                            highlightthickness=0, bd=0)
        _sb_canvas.pack(side=LEFT, fill=BOTH, expand=True)

        _sb_scroll = Scrollbar(scroll_outer, orient=VERTICAL,
                               command=_sb_canvas.yview)
        _sb_scroll.pack(side=RIGHT, fill=Y)
        _sb_canvas.configure(yscrollcommand=_sb_scroll.set)

        # inner frame that holds all buttons
        sb = Frame(_sb_canvas, bg=BG_SIDEBAR)
        _sb_win = _sb_canvas.create_window((0, 0), window=sb, anchor="nw")

        def _on_sb_configure(event):
            _sb_canvas.configure(scrollregion=_sb_canvas.bbox("all"))
        def _on_canvas_resize(event):
            _sb_canvas.itemconfig(_sb_win, width=event.width)

        sb.bind("<Configure>", _on_sb_configure)
        _sb_canvas.bind("<Configure>", _on_canvas_resize)

        # mouse-wheel scrolling — bind recursively to all children too
        def _on_mousewheel(event):
            _sb_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        def _on_linux_up(event):
            _sb_canvas.yview_scroll(-1, "units")
        def _on_linux_down(event):
            _sb_canvas.yview_scroll(1, "units")

        def _bind_mousewheel_recursive(widget):
            widget.bind("<MouseWheel>", _on_mousewheel)
            widget.bind("<Button-4>",   _on_linux_up)
            widget.bind("<Button-5>",   _on_linux_down)
            for child in widget.winfo_children():
                _bind_mousewheel_recursive(child)

        # Re-bind every time a new widget is added to sb
        def _on_sb_child_added(event):
            _bind_mousewheel_recursive(sb)

        _bind_mousewheel_recursive(_sb_canvas)
        _bind_mousewheel_recursive(sb)
        sb.bind("<Enter>", lambda e: _bind_mousewheel_recursive(sb))

        # ── Section: Students ──
        self._sidebar_section(sb, "STUDENTS")
        self._sidebar_btn(sb, "➕", "Add Student",        self.add_student,   ACCENT)
        self._sidebar_btn(sb, "🗑", "Delete Student",    self.delete_student, "#e67e22")
        self._sidebar_btn(sb, "⚡", "Train Model",       self.train_model,    ACCENT2)

        Frame(sb, bg=BORDER, height=1).pack(fill=X, padx=16, pady=6)
        self._sidebar_section(sb, "RECOGNITION")
        self._sidebar_btn(sb, "▶", "Start Recognition",  self.start_recognition, "#2ecc71")
        self._sidebar_btn(sb, "⏹", "Stop Recognition",  self.stop_recognition,  ACCENT_RED)

        Frame(sb, bg=BORDER, height=1).pack(fill=X, padx=16, pady=6)
        self._sidebar_section(sb, "EXPORT")
        self._sidebar_btn(sb, "📊", "Export Excel",      self.export_excel,   "#f39c12")
        self._sidebar_btn(sb, "📄", "Export PDF",        self.export_pdf,     "#9b59b6")

        Frame(sb, bg=BORDER, height=1).pack(fill=X, padx=16, pady=6)
        self._sidebar_section(sb, "SETTINGS")
        self._sidebar_btn(sb, "🔑", "Manage Users",      self.manage_users,   TEXT_DIM)
        self._sidebar_btn(sb, "📅", "Filter by Date",    self.filter_by_date, TEXT_DIM)
        self._sidebar_btn(sb, "↺",  "Reset Filter",      self.reset_filter,   TEXT_DIM)

        Frame(sb, bg=BORDER, height=1).pack(fill=X, padx=16, pady=6)
        self._sidebar_section(sb, "LOGS")
        self._sidebar_btn(sb, "🗂", "Saved Logs",       self.show_saved_logs,  "#00d4aa")

        Frame(sb, bg=BORDER, height=1).pack(fill=X, padx=16, pady=6)
        self._sidebar_btn(sb, "⏻",  "Exit",             self.exit_app,         ACCENT_RED)
        # Bind mousewheel to all widgets after all buttons are created
        self.master.after(100, lambda: _bind_mousewheel_recursive(sb))

        # ══════════════════ MAIN AREA ══════════════════
        main = Frame(self.master, bg=BG_DARK)
        main.pack(side=RIGHT, fill=BOTH, expand=True)

        # ── Top status bar ──
        topbar = Frame(main, bg="#0d1117", height=40)
        topbar.pack(fill=X, padx=0)
        topbar.pack_propagate(False)

        self.stat_present = Label(topbar,
            text="⬤  Present: 0", font=("Helvetica", 10, "bold"),
            bg=BG_DARK, fg="#2ecc71", padx=16)
        self.stat_present.pack(side=LEFT, pady=10)

        self.stat_unknown = Label(topbar,
            text="⬤  Unknown: 0", font=("Helvetica", 10),
            bg=BG_DARK, fg=ACCENT_RED, padx=8)
        self.stat_unknown.pack(side=LEFT)

        self.filter_label = Label(topbar, text="",
            font=("Helvetica", 9, "italic"), bg=BG_DARK, fg=TEXT_DIM)
        self.filter_label.pack(side=RIGHT, padx=16)

        # status indicator dot (live / idle)
        self._status_dot = Label(topbar, text="⬤  IDLE",
            font=("Helvetica", 9, "bold"), bg=BG_DARK, fg=TEXT_DIM, padx=16)
        self._status_dot.pack(side=RIGHT)

        # ── Camera preview card ──
        cam_card = Frame(main, bg=BG_CARD,
                         highlightthickness=2, highlightbackground=BORDER)
        cam_card.pack(padx=12, pady=(8, 6), fill=BOTH, expand=True)

        cam_header = Frame(cam_card, bg="#161b22")
        cam_header.pack(fill=X)
        Label(cam_header, text="  ● LIVE CAMERA FEED",
              font=("Helvetica", 9, "bold"), bg=BG_CARD, fg=ACCENT,
              padx=8, pady=6).pack(side=LEFT)
        Label(cam_header, text="LBPH Recognition Active",
              font=("Helvetica", 8), bg=BG_CARD, fg=TEXT_DIM,
              padx=8).pack(side=RIGHT)

        self._cam_card = cam_card   # keep ref for border animation

        self.preview_label = Label(cam_card, bg="#0a0a0f",
                                   text="▶  Press  「Start Recognition」  to activate camera",
                                   fg="#30363d", font=("Helvetica", 13))
        self.preview_label.pack(fill=BOTH, expand=True, padx=2, pady=2)

        # ── Attendance table card ──
        tbl_card = Frame(main, bg=BG_CARD,
                         highlightthickness=1, highlightbackground=BORDER)
        tbl_card.pack(padx=12, pady=(0, 6), fill=X)

        tbl_header = Frame(tbl_card, bg=BG_CARD)
        tbl_header.pack(fill=X)
        Label(tbl_header, text="  ◈ ATTENDANCE LOG",
              font=("Helvetica", 9, "bold"), bg=BG_CARD, fg=ACCENT2,
              padx=8, pady=6).pack(side=LEFT)

        # Save & Refresh buttons in log header
        def _make_log_btn(parent, text, cmd, color):
            b = Button(parent, text=text, font=("Helvetica", 8, "bold"),
                       bg=color, fg=BG_DARK, bd=0, relief="flat",
                       activebackground=color, cursor="hand2",
                       command=cmd, padx=10, pady=3)
            b.pack(side=RIGHT, padx=4, pady=4)
            return b

        _make_log_btn(tbl_header, "↺  Refresh", self.refresh_log,  "#e67e22")
        _make_log_btn(tbl_header, "💾  Save",   self.save_log,     ACCENT)

        tree_frame = Frame(tbl_card, bg=BG_CARD)
        tree_frame.pack(fill=X, padx=4, pady=(0, 4))

        cols = ("ID", "Name", "Date", "Time", "Status")
        self.tree = ttk.Treeview(tree_frame, columns=cols,
                                 show="headings", height=8,
                                 style="Dark.Treeview")
        widths = {"ID": 70, "Name": 180, "Date": 120, "Time": 100, "Status": 90}
        for col in cols:
            self.tree.heading(col, text=col.upper())
            self.tree.column(col, width=widths[col], anchor=CENTER)
        self.tree.tag_configure("present", background="#0d2818", foreground="#58d68d")
        self.tree.tag_configure("new_row", background="#1a3a4d")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                            command=self.tree.yview,
                            style="Dark.Vertical.TScrollbar")
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side=LEFT, fill=X, expand=True)
        vsb.pack(side=RIGHT, fill=Y)

        # Notification bar
        self.notif = NotificationBar(main)

        self.load_today_attendance()

    def _sidebar_section(self, parent, title):
        """Render a small uppercase section label in sidebar."""
        Label(parent, text=title, font=("Helvetica", 8, "bold"),
              bg=BG_SIDEBAR, fg=TEXT_DIM, anchor="w",
              padx=20).pack(fill=X, pady=(4, 2))

    def _sidebar_btn(self, parent, icon, label, command, accent_color):
        """Render a styled sidebar button with icon, label and hover glow."""
        row = Frame(parent, bg=BG_SIDEBAR, cursor="hand2")
        row.pack(fill=X, padx=12, pady=2)

        icon_lbl = Label(row, text=icon, width=2, font=("Helvetica", 11),
                         bg=BG_SIDEBAR, fg=accent_color)
        icon_lbl.pack(side=LEFT, padx=(6, 4))
        txt_lbl = Label(row, text=label, font=("Helvetica", 10),
                        bg=BG_SIDEBAR, fg=TEXT_PRIMARY, anchor="w")
        txt_lbl.pack(side=LEFT, fill=X, expand=True, ipady=6)

        accent_strip = Frame(row, bg=BG_SIDEBAR, width=3)
        accent_strip.pack(side=RIGHT, fill=Y)

        def on_enter(e):
            row.config(bg="#161b22")
            icon_lbl.config(bg="#161b22")
            txt_lbl.config(bg="#161b22", fg=accent_color)
            accent_strip.config(bg=accent_color)
        def on_leave(e):
            row.config(bg=BG_SIDEBAR)
            icon_lbl.config(bg=BG_SIDEBAR)
            txt_lbl.config(bg=BG_SIDEBAR, fg=TEXT_PRIMARY)
            accent_strip.config(bg=BG_SIDEBAR)
        def on_click(e):
            command()

        for w in (row, icon_lbl, txt_lbl):
            w.bind("<Enter>", on_enter)
            w.bind("<Leave>", on_leave)
            w.bind("<Button-1>", on_click)

    def _animate_cam_border(self):
        """Pulse the camera card border when recognition is running."""
        if self.running:
            self._cam_anim_step = (self._cam_anim_step + 1) % 40
            t = self._cam_anim_step / 40
            # oscillate between teal and dark
            g = int(170 + 85 * abs(t - 0.5) * 2)
            color = f"#00{min(g,255):02x}{'aa' if g > 200 else '66'}"
            self._cam_card.config(highlightbackground=color)
            self._status_dot.config(text="⬤  LIVE", fg="#2ecc71")
        else:
            self._cam_card.config(highlightbackground=BORDER)
            self._status_dot.config(text="⬤  IDLE", fg=TEXT_DIM)
        self.master.after(60, self._animate_cam_border)

    # ------------------------------------------------------------------ Helpers
    def typewriter_text(self, text, delay=80):
        self.current_text = ""
        self.full_text = text
        self.tw_index = 0
        def update():
            if self.tw_index < len(self.full_text):
                self.current_text += self.full_text[self.tw_index]
                self.header_label.config(text=self.current_text)
                self.tw_index += 1
                self.master.after(delay, update)
        update()

    def update_clock(self):
        now = datetime.datetime.now()
        self.clock_label.config(text=now.strftime("%a %d %b %Y\n%H : %M : %S"))
        self.master.after(1000, self.update_clock)

    def load_labels(self):
        labels = {}
        if os.path.exists(LABELS_FILE):
            with open(LABELS_FILE, "r") as f:
                for line in f:
                    parts = line.strip().split(":")
                    if len(parts) == 2:
                        labels[parts[0]] = parts[1]
        return labels

    def save_labels(self):
        with open(LABELS_FILE, "w") as f:
            for k, v in self.labels_map.items():
                f.write(f"{k}:{v}\n")

    def rgb_to_hex(self, rgb):
        return "#%02x%02x%02x" % rgb

    def flash_row(self, item, count=0):
        """Flash new row in teal then settle to present color."""
        if count > 6:
            self.tree.item(item, tags=("present",))
            return
        tag = "new_row" if count % 2 == 0 else "present"
        self.tree.item(item, tags=(tag,))
        self.master.after(200, lambda: self.flash_row(item, count + 1))

    def update_stats(self):
        present = len(self.tree.get_children())
        self.stat_present.config(text=f"⬤  Present: {present}")
        self.stat_unknown.config(text=f"⬤  Unknown: {self._unknown_count}")

    # ------------------------------------------------------------------ Attendance
    def load_today_attendance(self, date_filter=None):
        """Load attendance for a given date (default: today)."""
        self.tree.delete(*self.tree.get_children())
        target = date_filter or datetime.datetime.now().strftime("%Y-%m-%d")
        if os.path.exists(ATTENDANCE_FILE):
            with open(ATTENDANCE_FILE, "r", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    # Skip header row or any row where ID is not numeric
                    if not row or row[0].strip().upper() in ("ID", ""):
                        continue
                    if len(row) >= 4 and row[2] == target:
                        status = row[4] if len(row) >= 5 else "Present"
                        self.tree.insert("", "end",
                                         values=(row[0], row[1], row[2], row[3], status),
                                         tags=("present",))
        self.master.update_idletasks()
        self.update_stats()

    # ------------------------------------------------------------------ Student Mgmt
    def add_student(self):
        student_id = simpledialog.askstring("Add Student",
            "Enter Student ID (numbers only, e.g. 1, 101):", parent=self.master)
        if not student_id:
            return
        student_id = student_id.strip()

        # BUG FIX 1: ID must be a valid integer — train_model uses int(id) internally
        try:
            int(student_id)
        except ValueError:
            messagebox.showerror("Invalid ID",
                f"Student ID must be a number (e.g. 1, 42, 101).\n"
                f"You entered: '{student_id}'\n\nPlease try again.")
            return

        student_name = simpledialog.askstring("Add Student", "Enter Student Name:", parent=self.master)
        if not student_name:
            return

        # BUG FIX 2: Strip underscores from name — folder is split on '_' to extract ID
        student_name = student_name.strip().replace("_", " ")

        # Duplicate ID check — not allowed at all
        if student_id in self.labels_map:
            existing_name = self.labels_map[student_id]
            messagebox.showerror("Duplicate ID",
                f"Student ID  '{student_id}'  is already registered.\n\n"
                f"Existing student: {existing_name}\n\n"
                f"Please use a different ID, or delete the existing student first.")
            return

        # BUG FIX 3: Use a separator that won't collide — store as "ID_Name"
        # safe_name replaces spaces with hyphens to keep folder names clean
        safe_name = student_name.replace(" ", "-")
        save_dir = os.path.join(DATASET_DIR, f"{student_id}_{safe_name}")
        os.makedirs(save_dir, exist_ok=True)

        cap = cv2.VideoCapture(0)
        count = 0
        auto_capture = False

        # Paper §III-D: Support both manual (SPACE) and auto-capture modes
        info = messagebox.askyesno(
            "Capture Mode",
            f"Capture {IMAGES_PER_STUDENT} face images for {student_name}.\n\n"
            "YES = Auto-capture (continuous)\nNO = Manual (press SPACE each time)"
        )
        auto_capture = info
        messagebox.showinfo("Info",
            f"{'Auto' if auto_capture else 'Manual'} capture started.\n"
            "Look at the camera. Press ESC to cancel.")

        while count < IMAGES_PER_STUDENT:
            ret, frame = cap.read()
            if not ret:
                continue
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5, minSize=(80, 80))

            display = frame.copy()
            for (x, y, w, h) in faces:
                cv2.rectangle(display, (x, y), (x+w, y+h), (0, 255, 0), 2)

            progress_text = f"Captured: {count}/{IMAGES_PER_STUDENT}"
            cv2.putText(display, progress_text, (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 200, 255), 2)
            cv2.imshow("Capture Faces — ESC to cancel", display)

            key = cv2.waitKey(1 if auto_capture else 0) & 0xFF
            if key == 27:   # ESC
                break

            should_capture = (auto_capture and len(faces) > 0) or (key == 32 and len(faces) > 0)
            if should_capture:
                (x, y, w, h) = faces[0]
                face_img = cv2.resize(gray[y:y+h, x:x+w], (FACE_WIDTH, FACE_HEIGHT))
                # Paper §IV-I: Apply CLAHE preprocessing
                face_img = preprocess_face(face_img)
                cv2.imwrite(os.path.join(save_dir, f"{student_id}_{count}.png"), face_img)
                count += 1
                if auto_capture:
                    cv2.waitKey(150)  # short pause between auto captures

        cap.release()
        cv2.destroyAllWindows()

        if count > 0:
            self.labels_map[student_id] = student_name
            self.save_labels()
            self.notif.show(f"{student_name} added with {count} images. Train the model to activate.")
            messagebox.showinfo("Success", f"{student_name} added with {count} images.\nPlease click 'Train Model' to update recognition.")
        else:
            self.notif.error("No images captured. Student not added.")

    def delete_student(self):
        student_id = simpledialog.askstring("Delete Student",
                                            "Enter Student ID to delete:", parent=self.master)
        if not student_id:
            return
        if not messagebox.askyesno("Confirm Delete",
                f"Permanently delete student {student_id} and all their images?"):
            return
        deleted = False
        if os.path.exists(DATASET_DIR):
            for folder in os.listdir(DATASET_DIR):
                if folder.startswith(student_id + "_"):
                    import shutil
                    shutil.rmtree(os.path.join(DATASET_DIR, folder), ignore_errors=True)
                    deleted = True
        if student_id in self.labels_map:
            del self.labels_map[student_id]
            self.save_labels()
            deleted = True
        if deleted:
            self.notif.show(f"Student {student_id} deleted. Re-train model.")
            messagebox.showinfo("Deleted", f"Student {student_id} has been deleted.\nPlease re-train the model.")
        else:
            self.notif.error(f"Student {student_id} not found.")

    def train_model(self):
        """Train LBPH model. Paper §IV-E: accuracy improves with more images."""
        if not os.path.exists(DATASET_DIR) or not os.listdir(DATASET_DIR):
            messagebox.showwarning("No Data", "No student images found. Add students first.")
            return

        faces, ids = [], []
        student_counts = {}
        skipped_folders = []

        for folder in os.listdir(DATASET_DIR):
            path = os.path.join(DATASET_DIR, folder)
            if not os.path.isdir(path):
                continue

            # BUG FIX: Use split("_", 1) — name may contain underscores
            # Also strip whitespace that could cause int() to fail
            parts = folder.split("_", 1)
            raw_id = parts[0].strip()
            try:
                student_id = int(raw_id)
            except ValueError:
                # Folder doesn't start with a numeric ID — skip and warn
                skipped_folders.append(folder)
                continue

            img_count = 0
            for file in sorted(os.listdir(path)):
                if not file.lower().endswith((".png", ".jpg", ".jpeg")):
                    continue  # skip non-image files
                img_path = os.path.join(path, file)
                img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
                if img is not None:
                    img = cv2.resize(img, (FACE_WIDTH, FACE_HEIGHT))
                    img = preprocess_face(img)
                    faces.append(img)
                    ids.append(student_id)
                    img_count += 1
            if img_count:
                student_counts[folder] = img_count
            else:
                skipped_folders.append(f"{folder} (no readable images)")

        if not faces:
            skip_msg = ""
            if skipped_folders:
                skip_msg = "\n\nSkipped folders (bad format):\n" + "\n".join(f"  • {s}" for s in skipped_folders)
                skip_msg += "\n\nFolder names MUST start with a numeric ID, e.g.  '1_John'  or  '42_Alice'."
            messagebox.showwarning("No Valid Images",
                f"No images could be loaded for training.{skip_msg}\n\n"
                "Make sure you used a numeric Student ID when adding students.")
            return

        self.recognizer.train(faces, np.array(ids))
        self.recognizer.save(TRAINER_FILE)

        summary = "\n".join(f"  {k}: {v} images" for k, v in student_counts.items())
        skip_warn = ""
        if skipped_folders:
            skip_warn = "\n\n⚠ Skipped (non-numeric ID):\n" + "\n".join(f"  {s}" for s in skipped_folders)
        avg_imgs = len(faces) // max(len(student_counts), 1)
        self.notif.show(f"Model trained on {len(faces)} images across {len(student_counts)} students.")
        messagebox.showinfo("Training Complete",
            f"Model trained successfully!\n\nStudents trained:\n{summary}\n\n"
            f"Total images: {len(faces)}\n"
            f"Expected accuracy: {'~93%+' if avg_imgs >= 30 else '~85-90%' if avg_imgs >= 20 else '~80%'}"
            f"{skip_warn}")

    # ------------------------------------------------------------------ Recognition
    def start_recognition(self):
        if self.running:
            return
        if not os.path.exists(TRAINER_FILE):
            messagebox.showerror("Error", "No trained model found.\nPlease add students and train the model first.")
            return
        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            messagebox.showerror("Error", "Could not open webcam.")
            return
        self.running = True
        self._unknown_count = 0
        self._session_logged = set()     # new session — allow one log entry per person again
        self.notif.show("Recognition started.")
        self.recognize_loop()

    def recognize_loop(self):
        if not self.running or self.cap is None:
            return
        ret, frame = self.cap.read()
        if not ret:
            self.master.after(30, self.recognize_loop)
            return

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5,
                                              minSize=(80, 80))
        border_color = (40, 40, 40)

        for (x, y, w, h) in faces:
            face_img = cv2.resize(gray[y:y+h, x:x+w], (FACE_WIDTH, FACE_HEIGHT))
            face_img = preprocess_face(face_img)  # Paper §IV-I: CLAHE before recognition

            try:
                label_id, confidence = self.recognizer.predict(face_img)
            except Exception:
                continue

            # Paper §IV-B: confidence threshold 70-80; lower = better match
            if confidence <= RECOGNITION_CONFIDENCE_THRESHOLD:
                name = self.labels_map.get(str(label_id), "Unknown")
                color = (0, 200, 0)
                border_color = color
                conf_text = f"{name} ({100 - int(confidence)}%)"  # display as similarity %

                mark_attendance_once(label_id, name)  # save to CSV once per day

                # One log entry per person per start-stop session
                if str(label_id) not in self._session_logged:
                    self._session_logged.add(str(label_id))
                    now_str = datetime.datetime.now()
                    item = self.tree.insert("", "end",
                                            values=(label_id, name,
                                                    now_str.strftime("%Y-%m-%d"),
                                                    now_str.strftime("%H:%M:%S"),
                                                    "Present"),
                                            tags=("new_row",))
                    self.tree.see(item)
                    self.master.update_idletasks()
                    self.flash_row(item)
                    self._log_saved = False  # new row added, mark unsaved
                    self.update_stats()
                    self.notif.show(f"Detected: {name}")
            else:
                conf_text = f"Unknown ({100 - int(confidence)}%)"
                color = (0, 0, 220)
                border_color = color
                # Throttle unknown logging (once per 10s)
                now_ts = datetime.datetime.now().timestamp()
                face_key = f"{x//20}_{y//20}"
                if now_ts - self._unknown_log_cooldown.get(face_key, 0) > 10:
                    log_unknown_face()
                    self._unknown_log_cooldown[face_key] = now_ts
                    self._unknown_count += 1
                    self.update_stats()

            cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
            # Background for text readability
            (tw, th), _ = cv2.getTextSize(conf_text, cv2.FONT_HERSHEY_SIMPLEX, 0.65, 2)
            cv2.rectangle(frame, (x, y - th - 12), (x + tw + 6, y), color, -1)
            cv2.putText(frame, conf_text, (x + 3, y - 4),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)

        # Border color feedback
        self.preview_label.config(
            highlightbackground=self.rgb_to_hex(border_color),
            highlightcolor=self.rgb_to_hex(border_color),
            highlightthickness=4)

        # Display frame
        img_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        ph = self.preview_label.winfo_height() or 360
        pw = self.preview_label.winfo_width() or 520
        img_pil = Image.fromarray(img_rgb).resize((pw, ph))
        imgtk = ImageTk.PhotoImage(image=img_pil)
        self.preview_label.imgtk = imgtk
        self.preview_label.configure(image=imgtk, text="")

        self.master.after(30, self.recognize_loop)

    def stop_recognition(self):
        self.running = False
        if self.cap:
            self.cap.release()
            self.cap = None
        self.preview_label.config(image="",
                                  text="▶  Press  「Start Recognition」  to activate camera",
                                  fg="#30363d")
        # Wait 200ms for any in-flight recognize_loop + file writes to finish
        # before reloading the attendance log from CSV
        self.master.after(200, self._finish_stop)

    def _finish_stop(self):
        # Do NOT reload from CSV — keeps all live repeated log entries visible
        self.notif.show("Recognition stopped.")

    def exit_app(self):
        """Confirm and close the application cleanly."""
        if not self._log_saved and self.tree.get_children():
            ans = messagebox.askyesnocancel(
                "Unsaved Log",
                "You have unsaved log entries.\n\nSave before exiting?",
                parent=self.master)
            if ans is None:     # Cancel — abort exit
                return
            if ans:             # Yes — save first
                self.save_log()
                if not self._log_saved:
                    return      # user cancelled save dialog
        if messagebox.askokcancel("Exit", "Are you sure you want to exit?",
                                  parent=self.master):
            self.stop_recognition()
            self.master.after(250, self.master.destroy)

    # ------------------------------------------------------------------ Filtering
    def filter_by_date(self):
        """Paper §III-D: Attendance Reporting Panel — filter by date."""
        date_str = simpledialog.askstring(
            "Filter by Date", "Enter date (YYYY-MM-DD):",
            parent=self.master,
            initialvalue=datetime.datetime.now().strftime("%Y-%m-%d"))
        if date_str:
            try:
                datetime.datetime.strptime(date_str, "%Y-%m-%d")
                self.filter_date = date_str
                self.load_today_attendance(date_filter=date_str)
                self.filter_label.config(text=f"Showing: {date_str}")
                self.notif.show(f"Filtered to {date_str}")
            except ValueError:
                self.notif.error("Invalid date format. Use YYYY-MM-DD.")

    def reset_filter(self):
        self.filter_date = None
        self.filter_label.config(text="")
        self.load_today_attendance()
        self.notif.show("Showing today's attendance.")

    # ------------------------------------------------------------------ User Mgmt
    def manage_users(self):
        """Paper §III-D: secure login; allow adding/changing teacher accounts."""
        win = Toplevel(self.master)
        win.title("Manage Accounts")
        win.geometry("420x360")
        win.configure(bg=BG_DARK)
        win.resizable(False, False)
        win.grab_set()

        # header
        Frame(win, bg=ACCENT, height=3).pack(fill=X)
        Label(win, text="  ◈  MANAGE TEACHER ACCOUNTS",
              font=("Helvetica", 11, "bold"),
              bg=BG_CARD, fg=ACCENT, padx=12, pady=10).pack(fill=X)

        users_dict = load_users()

        frame = Frame(win, bg=BG_DARK)
        frame.pack(fill=BOTH, expand=True, padx=20, pady=12)

        Label(frame, text="CURRENT ACCOUNTS", font=("Helvetica", 8, "bold"),
              bg=BG_DARK, fg=TEXT_DIM).grid(row=0, column=0, sticky="w")
        lb = Listbox(frame, height=5, width=26, font=FONT_MONO,
                     bg=BG_CARD, fg=TEXT_PRIMARY, selectbackground=ACCENT,
                     selectforeground=BG_DARK, bd=0, highlightthickness=1,
                     highlightbackground=BORDER, relief="flat")
        lb.grid(row=1, column=0, columnspan=2, pady=6, sticky="ew")
        for u in users_dict:
            lb.insert(END, f"  {u}")

        Label(frame, text="USERNAME", font=("Helvetica", 8, "bold"),
              bg=BG_DARK, fg=TEXT_DIM).grid(row=2, column=0, sticky="w", pady=(8,2))
        new_user = Entry(frame, width=22, font=("Helvetica", 11),
                         bg="#21262d", fg=TEXT_PRIMARY, bd=0,
                         insertbackground=ACCENT, relief="flat",
                         highlightthickness=1, highlightbackground=BORDER,
                         highlightcolor=ACCENT)
        new_user.grid(row=3, column=0, columnspan=2, sticky="ew", ipady=6)

        Label(frame, text="PASSWORD", font=("Helvetica", 8, "bold"),
              bg=BG_DARK, fg=TEXT_DIM).grid(row=4, column=0, sticky="w", pady=(8,2))
        new_pass = Entry(frame, width=22, font=("Helvetica", 11), show="●",
                         bg="#21262d", fg=TEXT_PRIMARY, bd=0,
                         insertbackground=ACCENT, relief="flat",
                         highlightthickness=1, highlightbackground=BORDER,
                         highlightcolor=ACCENT)
        new_pass.grid(row=5, column=0, columnspan=2, sticky="ew", ipady=6)

        def add_user():
            u = new_user.get().strip()
            p = new_pass.get()
            if not u or not p:
                messagebox.showwarning("Input Error", "Both fields required.", parent=win)
                return
            users_dict[u] = hash_password(p)
            save_users(users_dict)
            lb.insert(END, f"  {u}")
            self.notif.show(f"User '{u}' saved.")

        def remove_user():
            sel = lb.curselection()
            if not sel:
                messagebox.showwarning("No Selection",
                    "Please select an account from the list first.", parent=win)
                return
            u = lb.get(sel[0]).strip()
            if u == "teacher" and len(users_dict) == 1:
                messagebox.showwarning("Cannot Remove",
                    "At least one admin account required.", parent=win)
                return
            if not messagebox.askyesno("Confirm Remove",
                    f"Remove account '{u}'?", parent=win):
                return
            if u not in users_dict:
                messagebox.showerror("Error",
                    f"Account '{u}' not found. Please try again.", parent=win)
                return
            del users_dict[u]
            save_users(users_dict)
            lb.delete(sel[0])
            self.notif.show(f"User '{u}' removed successfully.")

        btn_frame = Frame(frame, bg=BG_DARK)
        btn_frame.grid(row=6, column=0, columnspan=2, pady=14)
        for txt, cmd, color in [("Add / Update", add_user, ACCENT),
                                 ("Remove", remove_user, ACCENT_RED)]:
            b = Button(btn_frame, text=txt, font=("Helvetica", 10, "bold"),
                       bg=color, fg=BG_DARK, bd=0, relief="flat",
                       activebackground=color, cursor="hand2",
                       command=cmd, padx=14, pady=6)
            b.pack(side=LEFT, padx=6)

    # ------------------------------------------------------------------ Log Save/Refresh/View
    # ------------------------------------------------------------------ Session persistence
    def _load_sessions_from_disk(self):
        """Load saved sessions from JSON file on disk."""
        if not os.path.exists(SAVED_SESSIONS_FILE):
            return []
        try:
            with open(SAVED_SESSIONS_FILE, "r") as f:
                data = json.load(f)
            if isinstance(data, list):
                return data
        except Exception:
            pass
        return []

    def _persist_sessions(self):
        """Write current saved sessions to JSON file so they survive logout."""
        try:
            with open(SAVED_SESSIONS_FILE, "w") as f:
                json.dump(self._saved_sessions, f, indent=2)
        except Exception as e:
            self.notif.error(f"Could not persist sessions: {e}")

    def save_log(self):
        """Save current live log rows into a named session."""
        rows = []
        for iid in self.tree.get_children():
            rows.append(self.tree.item(iid)["values"])
        if not rows:
            messagebox.showwarning("Empty Log", "There are no log entries to save.", parent=self.master)
            return
        now_label = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        session_name = simpledialog.askstring(
            "Save Log", "Enter a name for this session log:",
            parent=self.master, initialvalue=f"Session {now_label}")
        if not session_name:
            return
        self._saved_sessions.append({"label": session_name, "rows": list(rows)})
        self._log_saved = True
        self._persist_sessions()
        self.notif.show(f"Log saved as: {session_name}")

    def refresh_log(self):
        """Clear the live log. If unsaved, ask the user first."""
        if not self.tree.get_children():
            self.notif.show("Log is already empty.")
            return
        if not self._log_saved:
            ans = messagebox.askyesnocancel(
                "Unsaved Log",
                "The current log has unsaved entries.\n\nDo you want to save before clearing?",
                parent=self.master)
            if ans is None:       # Cancel — do nothing
                return
            if ans:               # Yes — save first then clear
                self.save_log()
                if not self._log_saved:   # user cancelled the save dialog
                    return
        # Clear the live log and cooldowns so detections can re-appear fresh
        self.tree.delete(*self.tree.get_children())
        self._session_logged = set()     # reset so next session logs fresh
        self._log_saved = True
        self.update_stats()
        self.notif.show("Attendance log cleared.")

    def show_saved_logs(self):
        """Open a window listing all saved sessions; click one to view its entries."""
        win = Toplevel(self.master)
        win.title("Saved Attendance Logs")
        win.geometry("780x500")
        win.configure(bg=BG_DARK)
        win.grab_set()

        Frame(win, bg=ACCENT, height=3).pack(fill=X)
        Label(win, text="  🗂  SAVED ATTENDANCE LOGS",
              font=("Helvetica", 11, "bold"),
              bg=BG_CARD, fg=ACCENT, padx=12, pady=10).pack(fill=X)

        if not self._saved_sessions:
            Label(win, text="No saved sessions yet.\nUse the 💾 Save button on the Attendance Log to save a session.",
                  font=("Helvetica", 11), bg=BG_DARK, fg=TEXT_DIM,
                  justify=CENTER).pack(expand=True)
            return

        pane = Frame(win, bg=BG_DARK)
        pane.pack(fill=BOTH, expand=True, padx=12, pady=10)

        # Left: session list
        left = Frame(pane, bg=BG_CARD, width=220,
                     highlightthickness=1, highlightbackground=BORDER)
        left.pack(side=LEFT, fill=Y, padx=(0, 8))
        left.pack_propagate(False)

        Label(left, text="Sessions", font=("Helvetica", 9, "bold"),
              bg=BG_CARD, fg=TEXT_DIM, pady=6).pack(fill=X)

        lb = Listbox(left, font=FONT_MONO, bg=BG_CARD, fg=TEXT_PRIMARY,
                     selectbackground=ACCENT, selectforeground=BG_DARK,
                     bd=0, highlightthickness=0, relief="flat", activestyle="none")
        lb.pack(fill=BOTH, expand=True, padx=4, pady=(0, 4))
        for s in self._saved_sessions:
            lb.insert(END, f"  {s['label']}")

        # Right: treeview for selected session
        right = Frame(pane, bg=BG_CARD,
                      highlightthickness=1, highlightbackground=BORDER)
        right.pack(side=LEFT, fill=BOTH, expand=True)

        # Right header: title + row count + delete row button
        right_hdr = Frame(right, bg=BG_CARD)
        right_hdr.pack(fill=X)
        Label(right_hdr, text="Log Entries", font=("Helvetica", 9, "bold"),
              bg=BG_CARD, fg=TEXT_DIM, pady=6, padx=6).pack(side=LEFT)
        row_count_lbl = Label(right_hdr, text="", font=("Helvetica", 8),
                              bg=BG_CARD, fg=TEXT_DIM)
        row_count_lbl.pack(side=LEFT)

        cols = ("ID", "Name", "Date", "Time", "Status")
        detail_tree = ttk.Treeview(right, columns=cols, show="headings",
                                   height=15, style="Dark.Treeview")
        widths = {"ID": 60, "Name": 160, "Date": 110, "Time": 90, "Status": 80}
        for col in cols:
            detail_tree.heading(col, text=col)
            detail_tree.column(col, width=widths[col], anchor=CENTER)
        detail_tree.tag_configure("present", background="#0d2818", foreground="#58d68d")
        detail_tree.tag_configure("selected_row", background="#3a1a1a", foreground="#ff8080")

        vsb2 = ttk.Scrollbar(right, orient="vertical", command=detail_tree.yview,
                              style="Dark.Vertical.TScrollbar")
        detail_tree.configure(yscrollcommand=vsb2.set)
        detail_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(4, 0), pady=(0, 4))
        vsb2.pack(side=RIGHT, fill=Y, pady=(0, 4))

        def refresh_row_count():
            n = len(detail_tree.get_children())
            row_count_lbl.config(text=f"({n} entries)" if n else "")

        def on_select(event):
            sel = lb.curselection()
            if not sel:
                return
            detail_tree.delete(*detail_tree.get_children())
            session = self._saved_sessions[sel[0]]
            for row in session["rows"]:
                detail_tree.insert("", "end", values=row, tags=("present",))
            refresh_row_count()

        lb.bind("<<ListboxSelect>>", on_select)

        # ── Delete selected LOG ROW ──
        def delete_log_row():
            sel_rows = detail_tree.selection()
            if not sel_rows:
                messagebox.showwarning("No Selection",
                    "Please click on a log entry to select it first.", parent=win)
                return
            sess_sel = lb.curselection()
            if not sess_sel:
                return
            session = self._saved_sessions[sess_sel[0]]
            if not messagebox.askyesno("Delete Entry",
                    f"Delete {len(sel_rows)} selected log entr{'y' if len(sel_rows)==1 else 'ies'}?",
                    parent=win):
                return
            # Remove from in-memory session rows by matching values
            for iid in sel_rows:
                vals = list(detail_tree.item(iid)["values"])
                for i, r in enumerate(session["rows"]):
                    if list(r) == vals:
                        session["rows"].pop(i)
                        break
                detail_tree.delete(iid)
            refresh_row_count()
            self._persist_sessions()
            self.notif.show(f"Deleted {len(sel_rows)} log entr{'y' if len(sel_rows)==1 else 'ies'}.")

        Button(right_hdr, text="🗑  Delete Row",
               font=("Helvetica", 8, "bold"),
               bg=ACCENT_RED, fg="white", bd=0, relief="flat",
               activebackground=ACCENT_RED, cursor="hand2",
               command=delete_log_row, padx=10, pady=3).pack(side=RIGHT, padx=6, pady=4)

        # ── Bottom bar: delete session button ──
        def delete_session():
            sel = lb.curselection()
            if not sel:
                return
            if messagebox.askyesno("Delete Session",
                    f"Delete entire session '{self._saved_sessions[sel[0]]['label']}'?",
                    parent=win):
                self._saved_sessions.pop(sel[0])
                lb.delete(sel[0])
                detail_tree.delete(*detail_tree.get_children())
                refresh_row_count()
                self._persist_sessions()
                self.notif.show("Session deleted.")

        Button(win, text="🗑  Delete Entire Session",
               font=("Helvetica", 9, "bold"),
               bg="#8B0000", fg="white", bd=0, relief="flat",
               activebackground="#8B0000", cursor="hand2",
               command=delete_session, pady=6).pack(pady=(0, 8))

        # ------------------------------------------------------------------ Export helpers
    def _ask_export_date_filter(self, export_type):
        """
        Show a dialog asking the user to choose:
          1. By Date  — asks for a date and returns that date string
          2. Overall  — returns None (meaning all records)
        Returns (date_str_or_None, label_for_filename) or (False, False) if cancelled.
        """
        win = Toplevel(self.master)
        win.title(f"Export {export_type}")
        win.geometry("360x240")
        win.configure(bg=BG_DARK)
        win.resizable(False, False)
        win.grab_set()

        Frame(win, bg=ACCENT, height=3).pack(fill=X)
        Label(win, text=f"  ◈  EXPORT {export_type.upper()}",
              font=("Helvetica", 11, "bold"),
              bg=BG_CARD, fg=ACCENT, padx=12, pady=10).pack(fill=X)

        Label(win, text="Choose export type:", font=("Helvetica", 10),
              bg=BG_DARK, fg=TEXT_PRIMARY).pack(pady=(18, 8))

        result = {"value": False}  # False = cancelled

        def choose_by_date():
            win.destroy()
            date_str = simpledialog.askstring(
                "Enter Date", "Enter date (YYYY-MM-DD):",
                parent=self.master,
                initialvalue=datetime.datetime.now().strftime("%Y-%m-%d"))
            if not date_str:
                result["value"] = False
                return
            try:
                datetime.datetime.strptime(date_str, "%Y-%m-%d")
                result["value"] = ("date", date_str)
            except ValueError:
                messagebox.showerror("Invalid Date", "Use format YYYY-MM-DD")
                result["value"] = False

        def choose_overall():
            result["value"] = ("overall", None)
            win.destroy()

        def on_cancel():
            result["value"] = False
            win.destroy()

        btn_frame = Frame(win, bg=BG_DARK)
        btn_frame.pack(pady=10)

        for txt, cmd, color in [
            ("📅  By Date",    choose_by_date,  ACCENT),
            ("📋  Overall",    choose_overall,  ACCENT2),
        ]:
            Button(btn_frame, text=txt, font=("Helvetica", 11, "bold"),
                   bg=color, fg=BG_DARK, bd=0, relief="flat",
                   activebackground=color, cursor="hand2",
                   command=cmd, padx=18, pady=8).pack(side=LEFT, padx=10)

        Button(win, text="Cancel", font=("Helvetica", 9),
               bg=BG_CARD, fg=TEXT_DIM, bd=0, relief="flat",
               cursor="hand2", command=on_cancel).pack(pady=(4, 0))

        win.wait_window()
        return result["value"]

    def _read_attendance_rows(self, date_filter=None):
        """Read attendance CSV and return list of rows (excluding header).
           If date_filter is given, only rows matching that date are returned."""
        rows = []
        if not os.path.exists(ATTENDANCE_FILE):
            return rows
        with open(ATTENDANCE_FILE, "r", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                if not row or row[0].strip().upper() in ("ID", ""):
                    continue
                if date_filter and (len(row) < 3 or row[2] != date_filter):
                    continue
                rows.append(row)
        return rows

    # ------------------------------------------------------------------ Export
    def export_excel(self):
        """Export Excel — asks user: By Date or Overall."""
        if not os.path.exists(ATTENDANCE_FILE):
            messagebox.showwarning("Warning", "No attendance records to export.")
            return

        choice = self._ask_export_date_filter("Excel")
        if choice is False:
            return
        mode, date_filter = choice

        rows = self._read_attendance_rows(date_filter)
        if not rows:
            messagebox.showwarning("No Records",
                f"No attendance records found{' for ' + date_filter if date_filter else ''}.")
            return

        label = date_filter.replace("-", "") if date_filter else "Overall"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"Attendance_{label}.xlsx")
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name="Arial", size=11)

        # Write header
        headers = ["ID", "Name", "Date", "Time", "Status"]
        for c, val in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=val)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for r, row in enumerate(rows, 2):
            padded = row + [""] * (5 - len(row))
            for c, val in enumerate(padded[:5], 1):
                cell = ws.cell(row=r, column=c, value=val)
                if val == "Present":
                    cell.font = Font(color="2E7D32", bold=True)
                elif val == "Absent":
                    cell.font = Font(color="C62828", bold=True)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

        wb.save(file_path)
        self.notif.show(f"Excel exported: {os.path.basename(file_path)}")
        messagebox.showinfo("Success",
            f"Excel exported successfully!\n{file_path}\n\nTotal records: {len(rows)}")

    def export_pdf(self):
        """Export PDF — asks user: By Date or Overall."""
        if not os.path.exists(ATTENDANCE_FILE):
            messagebox.showwarning("Warning", "No attendance records to export.")
            return

        choice = self._ask_export_date_filter("PDF")
        if choice is False:
            return
        mode, date_filter = choice

        rows = self._read_attendance_rows(date_filter)
        if not rows:
            messagebox.showwarning("No Records",
                f"No attendance records found{' for ' + date_filter if date_filter else ''}.")
            return

        label = date_filter.replace("-", "") if date_filter else "Overall"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")],
            initialfile=f"AttendanceReport_{label}.pdf")
        if not file_path:
            return

        c_pdf = canvas.Canvas(file_path, pagesize=A4)
        width, height = A4

        title_text = f"Attendance Report — {date_filter}" if date_filter else "Attendance Report — All Records"

        c_pdf.setFillColorRGB(0.29, 0.56, 0.89)
        c_pdf.rect(0, height - 70, width, 70, fill=1, stroke=0)
        c_pdf.setFillColorRGB(1, 1, 1)
        c_pdf.setFont("Helvetica-Bold", 15)
        c_pdf.drawCentredString(width / 2, height - 28, title_text)
        c_pdf.setFont("Helvetica", 10)
        c_pdf.drawCentredString(width / 2, height - 50,
            f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}    |    Records: {len(rows)}")

        y = height - 90
        headers = ["ID", "Name", "Date", "Time", "Status"]
        col_x = [50, 100, 210, 320, 410]

        def draw_row(row_data, y_pos, is_header=False, status="Present"):
            if is_header:
                c_pdf.setFillColorRGB(0.29, 0.56, 0.89)
                c_pdf.rect(45, y_pos - 4, width - 90, 18, fill=1, stroke=0)
                c_pdf.setFillColorRGB(1, 1, 1)
                c_pdf.setFont("Helvetica-Bold", 9)
            else:
                if status == "Present":
                    c_pdf.setFillColorRGB(0.91, 0.97, 0.91)
                else:
                    c_pdf.setFillColorRGB(1, 0.93, 0.93)
                c_pdf.rect(45, y_pos - 4, width - 90, 18, fill=1, stroke=0)
                c_pdf.setFillColorRGB(0, 0, 0)
                c_pdf.setFont("Helvetica", 9)
            for i, val in enumerate(row_data):
                c_pdf.drawString(col_x[i], y_pos, str(val)[:18])

        draw_row(headers, y, is_header=True)
        y -= 22

        for row in rows:
            if y < 60:
                c_pdf.showPage()
                y = height - 60
                draw_row(headers, y, is_header=True)
                y -= 22
            status = row[4] if len(row) >= 5 else "Present"
            padded = row + [""] * (5 - len(row))
            draw_row(padded[:5], y, status=status)
            y -= 20

        c_pdf.save()
        self.notif.show(f"PDF exported: {os.path.basename(file_path)}")
        messagebox.showinfo("Success", f"PDF report exported successfully!\n{file_path}")


# -------------------- Main --------------------
def main_app():
    root = Tk()
    app = AttendanceApp(root)

    def on_close():
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            app.stop_recognition()
            root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


if __name__ == "__main__":
    login_root = Tk()
    LoginWindow(login_root, on_success=main_app)
    login_root.mainloop()
